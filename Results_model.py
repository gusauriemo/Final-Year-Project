import nengo
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from scipy.io import loadmat
import itertools
import pandas as pd
from openpyxl import load_workbook
import scipy.io
from subset import subsets

def model_function(sim_runtime1, epoch, test_loss= False):
    from error import hearing_prediction, correct_prediction, detection_error_loss, distance_loss, sethares
    #File Paths
    full_path = "matrix path" #i.e., the full dataset
    labels = "matrix path"

    #Handling the data
    matrix_data = loadmat(full_path)
    label_data = loadmat(labels)
    input_array = (np.log10(matrix_data["matrix"]))
    input_array[input_array<0] = 0
    output_array = (label_data["labels"])*3

    #Global variables
    nn= 30 #number of notes
    initial = 0
    test_loss_index = round((sim_runtime1/(0.15)))
    final = initial+nn
    D = nn #dimensions
    N = D*50 #base number of neurons

    decoder_path = "Decoder path"+str(epoch)+ "/decoders_30_"+str(sim_runtime1)+"_"+ str(epoch)+".mat"
    decoder_data = loadmat(decoder_path)
    decoders1 = decoder_data["decoders"]
    decoders = decoders1[0,:,:]

    if test_loss == True:
        sim_runtime = round(sim_runtime1*(0.3))
    else:
        sim_runtime = sim_runtime1

    new_indexes = subsets(output_array, nn)

    #check if this is a training loss simulation or a test loss simulation
    if test_loss == True:
        test_index = round(sim_runtime1*6.6)
    else:
        test_index=0

    with nengo.Network(seed=0) as net:

        # Desired output
        filtered_output_array = output_array[new_indexes, initial:final]
        correct_stim = nengo.Node(nengo.processes.PresentInput(filtered_output_array[test_index:(len(filtered_output_array)),:], presentation_time=0.15))
        print(len(filtered_output_array[test_index:(len(filtered_output_array)),:]))
        # Hearing
        filtered_input_array = input_array[new_indexes, initial:final]
        hearing_stim = nengo.Node(nengo.processes.PresentInput(filtered_input_array[test_index:(len(filtered_output_array)),:], presentation_time=0.15))
        hearing_ensemble = nengo.Ensemble(N*4, D, radius=4, seed=0)
        decision_ensemble = nengo.Ensemble(N*5, D, radius=4, seed=0)

        #Connections
        nengo.Connection(hearing_stim, hearing_ensemble[:D])

        #Direct neurons to ensemble connection
        conn2 = nengo.Connection(hearing_ensemble.neurons, decision_ensemble, transform=decoders)

        #Probes
        probe1 = nengo.Probe(hearing_ensemble,  synapse=0.01)
        probe4 = nengo.Probe(decision_ensemble, synapse=0.01) 
        probe2 = nengo.Probe(decision_ensemble, synapse=0.01, sample_every=0.15) 
        """this probe gives a matrix of the output values every 0.15s. 
        Currently, since only 30 notes are being input, the number of columns is 30. 
        The number of rows will be determined by sim_runtime/0.15 (to the lower bound)"""
        probe3 = nengo.Probe(correct_stim, synapse=0.01)
        spike_probe = nengo.Probe(decision_ensemble.neurons)
        spike_probe2 = nengo.Probe(hearing_ensemble.neurons)

    with nengo.Simulator(net) as sim:
        sim.run(sim_runtime)

        spike_count = np.sum(sim.data[spike_probe] > 0, axis=0) #total number of spikes in decision_ensemble during simulation
        spike_count2 = np.sum(sim.data[spike_probe2] > 0, axis=0) #total number of spikes in hearing_ensemble

    #Model Characteristics
    number_spikes = np.sum(spike_count) + np.sum(spike_count2)
    number_neurons = net.n_neurons
    print(f"Total numbers of spikes: {number_spikes}")
    print(f"Total number of neurons: {number_neurons}")
    print(f"Average spikes per neuron per second: {np.round((number_spikes/number_neurons)/sim_runtime, decimals=2)}")


    #Error Calculations
    last= round((sim_runtime/(0.15))-1) #number of error points
    hearing_sense = (hearing_prediction(sim.data[probe2]))[0:last]
    comparison = new_indexes[0:last]
    expectation = (correct_prediction(output_array[comparison]))

    de_loss = (np.mean(detection_error_loss(expectation, hearing_sense)))
    distance_loss = (distance_loss(expectation, hearing_sense))
    #distance_loss1 = np.mean(distance_loss(expectation, hearing_sense))
    consonance_loss = np.mean(sethares(expectation, hearing_sense)[0])
    consonance_loss1 = np.mean(sethares(expectation, hearing_sense)[1])

    #Result Outputs

    print(f"The accuracy under detection-error loss is {np.round((1-de_loss)*100, decimals= 1)}%")
    print(f"The accuracy under distance loss is {np.round((1-(np.mean(distance_loss[0])))*100, decimals=1)}%")
    print(f"The accuracy under linear distance loss is {np.round((1-(np.mean(distance_loss[1])))*100, decimals=1)}%")
    print(f"The accuracy under consonance loss is {np.round((1-consonance_loss)*100, decimals=1)}%")
    print(f"The accuracy under linear consonance loss is {np.round((1-consonance_loss1)*100, decimals=1)}%")


    #Excel appending
    data = {"runtime":[sim_runtime],
            "error_points":[last], 
            "DTE": [np.round((1-de_loss)*100, decimals= 1)],
            "DL":[np.round((1-(np.mean(distance_loss[0])))*100, decimals=1)],
            "LDL": [np.round((1-(np.mean(distance_loss[1])))*100, decimals=1)],
            "CL": [np.round((1-consonance_loss)*100, decimals=1)],
            "LCL": [np.round((1-consonance_loss1)*100, decimals=1)],
            "decoder": decoder_path,
            "Average spikes" : [np.round((number_spikes/number_neurons)/sim_runtime, decimals=2)],
            "Test loss" : [str(test_loss)],
            "Epoch" :[epoch]
            }
    df1 = pd.DataFrame(data)

    if test_loss == True:
        filename = 'test loss excel path'
    else:
        filename = 'training loss excel path'

    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    df1.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)

    writer.save() #https://stackoverflow.com/questions/47737220/append-dataframe-to-excel-with-pandas/47738103